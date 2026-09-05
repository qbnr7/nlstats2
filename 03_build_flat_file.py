#!/usr/bin/env python3
"""
03_build_flat_file.py
=====================
Reads all game Excel files and the schedule file, then produces a single
flat CSV file with one row per graded official call.

Each penalty on a play that involves multiple officials (e.g. LCHC) is
split into individual rows - one per official. Plays with two penalties
(PENALTY-CAT 1 and PENALTY CAT 2) are both processed.

Input:
    data/           - folder with one Excel file per game
                      filename must match the GameID in the schedule
    nlplan/         - folder with the schedule Excel file
                      requires sheets: 'Plan - NL' and 'Officials and games'

Output:
    output/flat_calls.csv  - one row per graded official call

Usage:
    python 03_build_flat_file.py

Changelog:
    v2 - H (legacy "Head Linesman" position code) is now recognised and
         normalised to D ("Down Judge") everywhere - in schedule columns
         and in GRADE OFFICIAL strings. D is the dominant/current code:
         if a schedule row somehow has both an H and a D value, D wins.
"""

import pandas as pd
import zipfile
import io
from pathlib import Path
from openpyxl import load_workbook

# ── Configuration ─────────────────────────────────────────────────────────────

DATA_FOLDER     = Path("data")
SCHEDULE_FOLDER = Path("nlplan")
OUTPUT_FOLDER   = Path("output")

SCHEDULE_SHEET  = "Plan - NL"
OFFICIALS_SHEET = "Officials and games"

# 'H' (Head Linesman) is the legacy code for the position now called
# 'D' (Down Judge). Both are accepted as input; 'H' is normalised to 'D'
# immediately so nothing downstream ever sees an 'H' again. If a game
# somehow has both an H and a D value, D wins (see POSITION_READ_ORDER,
# which reads H first so a later D value overwrites it).
POSITION_CODES      = {'R', 'U', 'D', 'H', 'L', 'B', 'F', 'S', 'C'}
POSITION_READ_ORDER = ['R', 'U', 'H', 'D', 'L', 'B', 'F', 'S', 'C']
GRADE_CODES          = {'C', 'M', 'I', 'N', 'G', 'W'}

PENALTY_COLUMNS = [
    ('PENALTY-CAT 1', 'FLAG 1', 'GRADE OFFICIAL 1'),
    ('PENALTY CAT 2',  'FLAG 2', 'GRADE OFFICIAL 2'),
]

# Minimal valid styles.xml - injected into xlsx files that are missing it
MINIMAL_STYLES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts><font><sz val="11"/><name val="Calibri"/></font></fonts>
  <fills>
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
  </fills>
  <borders><border><left/><right/><top/><bottom/><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>
</styleSheet>"""


# Counts how many times the legacy 'H' position code was seen and
# silently normalised, across schedule columns and GRADE OFFICIAL codes.
# Printed as a one-line summary at the end of main() so it doesn't get
# lost in the per-occurrence warnings below.
_legacy_h_count = 0


def normalise_position(pos, context=''):
    """
    Map legacy position code H (Head Linesman) to D (Down Judge).
    H should never appear in current data -- prints a warning whenever it
    is encountered so stale files get noticed and fixed at the source,
    but still normalises to D so processing is not interrupted.
    """
    global _legacy_h_count
    if pos == 'H':
        _legacy_h_count += 1
        print(f"    WARNING: legacy position code 'H' found "
              f"{context} -- normalising to 'D'. 'H' should not appear "
              f"in current data; check the source file.")
        return 'D'
    return pos

# ── Excel reader ───────────────────────────────────────────────────────────────

def load_xlsx(file_path, sheet_name=None):
    """
    Load an xlsx file, injecting a minimal styles.xml if the file is missing
    one. Returns a pandas DataFrame. If sheet_name is None, reads the first
    sheet.
    """
    with zipfile.ZipFile(file_path, 'r') as zin:
        names = zin.namelist()
        buf = io.BytesIO()

        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in names:
                if item == '[Content_Types].xml' and 'xl/styles.xml' not in names:
                    content = zin.read(item).decode()
                    content = content.replace(
                        '</Types>',
                        '<Override PartName="/xl/styles.xml" '
                        'ContentType="application/vnd.openxmlformats-officedocument'
                        '.spreadsheetml.styles+xml"/></Types>'
                    )
                    zout.writestr(item, content)
                else:
                    zout.writestr(item, zin.read(item))

            if 'xl/styles.xml' not in names:
                zout.writestr('xl/styles.xml', MINIMAL_STYLES)

    buf.seek(0)
    try:
        wb = load_workbook(buf, read_only=True, data_only=True)
    except Exception:
        # styles.xml was present but malformed (e.g. an empty <fill></fill>
        # with no patternFill/gradientFill, which some Excel exports
        # produce and which openpyxl refuses to parse). Retry with the
        # broken styles.xml swapped out for the minimal one.
        buf = io.BytesIO()
        with zipfile.ZipFile(file_path, 'r') as zin:
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.namelist():
                    if item == 'xl/styles.xml':
                        zout.writestr(item, MINIMAL_STYLES)
                    else:
                        zout.writestr(item, zin.read(item))
        buf.seek(0)
        wb = load_workbook(buf, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. "
                             f"Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    data = [[cell.value for cell in row] for row in ws.rows]
    wb.close()

    if not data:
        return pd.DataFrame()

    # openpyxl's read-only row iterator trims each row to its own last
    # non-empty cell, so rows can come back shorter than the header row
    # whenever trailing columns (e.g. notes columns) are sparsely filled.
    # Pad every row out to the widest row in the sheet so all rows line
    # up with the header before building the DataFrame.
    max_len = max(len(row) for row in data)
    data = [row + [None] * (max_len - len(row)) for row in data]

    headers = [str(h).strip() if h is not None else '' for h in data[0]]
    return pd.DataFrame(data[1:], columns=headers)

# ── Grade official parser ──────────────────────────────────────────────────────

def parse_grade_official(code, context=''):
    """
    Parse a GRADE OFFICIAL string into a list of (position, grade) pairs.
    Legacy position code H is normalised to D before being returned.

    e.g. 'RC' -> [('R','C')], 'LCDC' -> [('L','C'),('D','C')]
         'LCHC' -> [('L','C'),('D','C')]   (H normalised to D)
    """
    if pd.isna(code) or str(code).strip() == '':
        return []

    code = str(code).strip().upper()
    pairs = []
    i = 0

    while i < len(code) - 1:
        pos   = code[i]
        grade = code[i + 1]

        if pos in POSITION_CODES and grade in GRADE_CODES:
            prefix = f"in {context}, " if context else "in "
            full_context = f"{prefix}GRADE OFFICIAL code '{code}'"
            pairs.append((normalise_position(pos, context=full_context), grade))
            i += 2
        else:
            print(f"    Warning: unexpected characters '{code[i:i+2]}' "
                  f"in '{code}', skipping")
            i += 1

    return pairs

# ── Data loaders ──────────────────────────────────────────────────────────────

def load_officials(schedule_file):
    """
    Load officials database. Returns dict: { initials -> full name }
    """
    df = load_xlsx(schedule_file, sheet_name=OFFICIALS_SHEET)

    # In the raw sheet, row 1 is a merged title row and row 2 has the
    # real column headers ("Initialer", "Navn", ...), with official data
    # starting at row 3. But load_xlsx() already consumes the sheet's
    # first row (the blank title row) as the DataFrame's own column
    # headers, so by the time it gets here df.iloc[0] is already the
    # real header row -- not df.iloc[1]. Using iloc[1]/df[2:] here (as
    # if df still had the raw, unshifted rows) skips the true header row
    # and turns the first official's own data into bogus column names.
    df.columns = [str(c).strip() for c in df.iloc[0]]
    df = df[1:].reset_index(drop=True)

    officials = {}
    for _, row in df.iterrows():
        initials = str(row.get('Initialer', '') or '').strip()
        name     = str(row.get('Navn',      '') or '').strip()
        if initials and name and initials.lower() != 'nan' and name.lower() != 'nan':
            officials[initials] = name

    print(f"  Loaded {len(officials)} officials")
    return officials


def load_schedule(schedule_file):
    """
    Load game schedule. Returns dict: { game_id -> { date, home_team,
    away_team, positions } }

    Positions are read in POSITION_READ_ORDER (H before D) and normalised,
    so if a schedule row has both an H and a D value for the same
    position, D (the current/dominant code) wins.
    """
    df = load_xlsx(schedule_file, sheet_name=SCHEDULE_SHEET)

    schedule = {}
    for _, row in df.iterrows():
        game_id = str(row.get('GameID', '') or '').strip()
        if not game_id or game_id.lower() == 'nan':
            continue

        dato  = str(row.get('Dato',  '') or '').strip()
        maned = str(row.get('Måned', '') or '').strip()
        date  = f"{dato}-{maned}" if dato and maned else ''

        positions = {}
        for pos in POSITION_READ_ORDER:
            val = str(row.get(pos, '') or '').strip()
            if val and val.lower() != 'nan':
                val = val.split('+')[0].strip()
                # Strip parens only if the whole value is a backup
                # marker: (BT) -> BT, but keep nationality suffixes
                # like FK(DE) as-is. Must match 04_generate_reports.py's
                # load_schedule() exactly -- otherwise the same official
                # ends up keyed as both "(BT)" and "BT" across the two
                # scripts, splitting one person's calls and games
                # between two separate, unmatched identities.
                if val.startswith('(') and val.endswith(')'):
                    val = val[1:-1].strip()
                # H is read before D, so a D column (if present)
                # overwrites it -- D is dominant.
                norm_pos = normalise_position(
                    pos, context=f"in schedule column '{pos}' for game '{game_id}'")
                positions[norm_pos] = val

        schedule[game_id] = {
            'date':      date,
            'home_team': str(row.get('Hjemme', '') or '').strip(),
            'away_team': str(row.get('Ude',    '') or '').strip(),
            'positions': positions,
        }

    print(f"  Loaded {len(schedule)} games from schedule")
    return schedule

# ── Game file processor ────────────────────────────────────────────────────────

def clean_numeric(val):
    """
    Format a PLAY # / QTR value cleanly.

    A genuinely empty cell comes back from pandas as NaN, not None, once
    it shares a column with real values -- str(NaN) is the literal text
    'nan', not ''. And a numeric column that has any missing cells gets
    promoted to float64 by pandas (it has no integer "missing" value),
    so a whole-number play/quarter can come through as e.g. 1.0 instead
    of 1. This returns '' for missing values and strips a spurious .0
    from whole numbers, instead of passing either artifact straight
    into the output.
    """
    if pd.isna(val):
        return ''
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def process_game_file(file_path, game_id, game_info, officials):
    """
    Process a single game Excel file.
    Returns list of flat row dicts, one per graded official call.
    """
    df = load_xlsx(file_path)
    rows = []

    for _, play in df.iterrows():
        play_number = clean_numeric(play.get('PLAY #', ''))
        qtr         = clean_numeric(play.get('QTR',    ''))

        for penalty_col, flag_col, grade_col in PENALTY_COLUMNS:
            foul_code      = play.get(penalty_col, '')
            flag           = play.get(flag_col,    '')
            grade_official = play.get(grade_col,   '')

            if pd.isna(foul_code) or str(foul_code).strip() == '':
                continue

            foul_code = str(foul_code).strip()
            flag      = '' if pd.isna(flag) else str(flag).strip()
            pairs     = parse_grade_official(
                grade_official,
                context=f"game '{game_id}', play {play_number}")

            if not pairs:
                rows.append(build_row(game_id, game_info, play_number, qtr,
                                      foul_code, flag, '', '', '', ''))
                continue

            for position, grade in pairs:
                initials = game_info['positions'].get(position, '')
                name     = officials.get(initials, '') if initials else ''
                rows.append(build_row(game_id, game_info, play_number, qtr,
                                      foul_code, flag, position, initials,
                                      name, grade))
    return rows


def build_row(game_id, game_info, play_number, qtr,
              foul_code, flag, position, initials, name, grade):
    return {
        'game_id':           game_id,
        'date':              game_info['date'],
        'home_team':         game_info['home_team'],
        'away_team':         game_info['away_team'],
        'play_number':       play_number,
        'qtr':               qtr,
        'foul_code':         foul_code,
        'flag':              flag,
        'position':          position,
        'official_initials': initials,
        'official_name':     name,
        'grade_code':        grade,
    }

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("03_build_flat_file.py")
    print("=" * 50)

    for folder in [DATA_FOLDER, SCHEDULE_FOLDER]:
        if not folder.exists():
            print(f"ERROR: Folder '{folder}' not found")
            return

    OUTPUT_FOLDER.mkdir(exist_ok=True)

    schedule_files = list(SCHEDULE_FOLDER.glob("*.xlsx")) + \
                     list(SCHEDULE_FOLDER.glob("*.xls"))
    if not schedule_files:
        print(f"ERROR: No Excel files found in '{SCHEDULE_FOLDER}'")
        return
    schedule_file = schedule_files[0]
    print(f"\nSchedule file : {schedule_file.name}")

    officials = load_officials(schedule_file)
    schedule  = load_schedule(schedule_file)

    game_files = sorted(
        list(DATA_FOLDER.glob("*.xlsx")) + list(DATA_FOLDER.glob("*.xls"))
    )
    if not game_files:
        print(f"ERROR: No Excel files found in '{DATA_FOLDER}'")
        return
    print(f"  Found {len(game_files)} game file(s)\n")

    all_rows  = []
    matched   = 0
    unmatched = []

    for file_path in game_files:
        game_id = file_path.stem

        if game_id in schedule:
            game_info = schedule[game_id]
            matched += 1
        else:
            print(f"  Warning: '{game_id}' not found in schedule")
            game_info = {'date': '', 'home_team': '', 'away_team': '',
                         'positions': {}}
            unmatched.append(game_id)

        print(f"  Processing : {file_path.name}")
        rows = process_game_file(file_path, game_id, game_info, officials)
        all_rows.extend(rows)
        print(f"  Rows output: {len(rows)}\n")

    if not all_rows:
        print("No graded calls found - nothing written")
        return

    output_path = OUTPUT_FOLDER / "flat_calls.csv"
    pd.DataFrame(all_rows, columns=[
        'game_id', 'date', 'home_team', 'away_team',
        'play_number', 'qtr', 'foul_code', 'flag',
        'position', 'official_initials', 'official_name', 'grade_code'
    ]).to_csv(output_path, index=False)

    print("-" * 50)
    print(f"Output file   : {output_path}")
    print(f"Total rows    : {len(all_rows)}")
    print(f"Games matched : {matched}")
    if unmatched:
        print(f"Not in schedule ({len(unmatched)}):")
        for g in unmatched:
            print(f"  - {g}")
    if _legacy_h_count:
        print(f"\nWARNING: legacy position code 'H' was found and "
              f"normalised to 'D' {_legacy_h_count} time(s) this run -- "
              f"see warnings above for exactly where. 'H' should not "
              f"appear in current data; the source schedule/game files "
              f"should be updated to use 'D' instead.")


if __name__ == "__main__":
    main()
